import tkinter as tk
from ttkbootstrap import Style
from ttkbootstrap.widgets import (
    LabelFrame, Frame, Button, Label, Entry, Combobox, Spinbox, Checkbutton
)
from tkinter import filedialog, colorchooser, font, messagebox
import openpyxl
import json
import os
import difflib

SETTINGS_FILE = "excel_comparator_settings.json"
RECENT_LIMIT = 10

def safe_color(color):
    if not color: color = "#FFFFFF"
    if color.startswith("#"): color = color[1:]
    if len(color) == 6: color = "FF" + color.upper()
    if len(color) == 8: color = color.upper()
    else: color = "FFFFFFFF"
    return color

def get_fill(color):
    from openpyxl.styles import PatternFill
    return PatternFill(fill_type="solid", fgColor=safe_color(color))

def get_font(family, size, bold=False, color="#000000"):
    from openpyxl.styles import Font
    return Font(name=family, size=size, bold=bold, color=safe_color(color))

def get_border(thickness, color):
    from openpyxl.styles import Border, Side
    border_style = {0: None, 1: "thin", 2: "medium", 3: "thick"}.get(thickness, "thin")
    side = Side(border_style=border_style, color=safe_color(color))
    return openpyxl.styles.Border(left=side, right=side, top=side, bottom=side)

def autofit_columns(ws, extra_padding=2):
    for col in ws.columns:
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(val))
        ws.column_dimensions[col_letter].width = min(50, max(8, max_length + extra_padding))

def mapping_str_to_int(d):
    return {int(k): int(v) for k, v in d.items()}

def suggest_mappings(headers1, headers2):
    mapping = {}
    headers2_lower = [h.lower() for h in headers2]
    for i, h1 in enumerate(headers1):
        h1_lower = str(h1).lower()
        if h1_lower in headers2_lower:
            mapping[i] = headers2_lower.index(h1_lower)
        else:
            matches = difflib.get_close_matches(h1_lower, headers2_lower, n=1, cutoff=0.8)
            if matches and matches[0] in headers2_lower:
                mapping[i] = headers2_lower.index(matches[0])
    return mapping

class MappingDialog(tk.Toplevel):
    def __init__(self, master, headers1, headers2, mapdict, include1, include2):
        super().__init__(master)
        self.title("Map Columns (File 1 → File 2)")
        self.geometry("950x600")
        self.minsize(700, 320)
        self.resizable(True, True)
        self.headers1 = headers1
        self.headers2 = headers2
        self.mapdict = dict(mapdict)
        self.include1 = list(include1)
        self.include2 = list(include2)
        self.result = None
        self.result_include1 = None
        self.result_include2 = None

        profile_frame = Frame(self)
        profile_frame.pack(fill="x", side="top", pady=(0,2))
        Button(profile_frame, text="Save Mapping…", command=self.save_mapping_profile, bootstyle="secondary-outline").pack(side="left", padx=(16,8), pady=(6,2))
        Button(profile_frame, text="Load Mapping…", command=self.load_mapping_profile, bootstyle="secondary-outline").pack(side="left", padx=(2,8), pady=(6,2))

        outer_frame = Frame(self)
        outer_frame.pack(fill="both", expand=True)
        canvas = tk.Canvas(outer_frame, borderwidth=0)
        vscroll = tk.Scrollbar(outer_frame, orient="vertical", command=canvas.yview)
        scroll_frame = Frame(canvas)

        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=vscroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        vscroll.pack(side="right", fill="y")

        sel_frame = Frame(scroll_frame)
        sel_frame.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0,8))
        Label(sel_frame, text="Include columns in comparison:").pack(side="left", padx=2)
        self.chk_vars1 = []
        for i, h in enumerate(headers1):
            v = tk.BooleanVar(value=include1[i])
            c = Checkbutton(sel_frame, text=f"1:{h}", variable=v, bootstyle="success-round-toggle")
            c.pack(side="left", padx=(0,2))
            self.chk_vars1.append(v)
        Label(sel_frame, text="|").pack(side="left")
        self.chk_vars2 = []
        for i, h in enumerate(headers2):
            v = tk.BooleanVar(value=include2[i])
            c = Checkbutton(sel_frame, text=f"2:{h}", variable=v, bootstyle="info-round-toggle")
            c.pack(side="left", padx=(0,2))
            self.chk_vars2.append(v)

        l1 = tk.Listbox(scroll_frame, exportselection=False, width=28, height=15)
        l2 = tk.Listbox(scroll_frame, exportselection=False, width=28, height=15)
        for h in headers1: l1.insert(tk.END, h)
        for h in headers2: l2.insert(tk.END, h)
        l1.grid(row=1, column=0, rowspan=2, padx=10, pady=10, sticky="ns")
        l2.grid(row=1, column=2, rowspan=2, padx=10, pady=10, sticky="ns")
        self.l1 = l1
        self.l2 = l2

        btn_frame = Frame(scroll_frame)
        btn_frame.grid(row=1, column=1, rowspan=2, sticky="ns")
        map_btn = Button(btn_frame, text="Map →", command=self.map_selected, bootstyle="success-outline")
        unmap_btn = Button(btn_frame, text="Unmap", command=self.unmap_selected, bootstyle="danger-outline")
        map_btn.pack(pady=10)
        unmap_btn.pack(pady=10)
        suggest_btn = Button(btn_frame, text="Suggest", command=self.suggest, bootstyle="info-outline")
        suggest_btn.pack(pady=10)
        self.mapping_view = tk.Listbox(scroll_frame, width=48)
        self.mapping_view.grid(row=3, column=0, columnspan=3, pady=8, sticky="ew")
        self.mapping_view.bind('<Double-1>', lambda e: self.unmap_selected())
        scroll_frame.grid_columnconfigure(0, weight=1)
        scroll_frame.grid_columnconfigure(1, weight=0)
        scroll_frame.grid_columnconfigure(2, weight=1)

        bottom_frame = Frame(self)
        bottom_frame.pack(fill="x", side="bottom", pady=(0,12))
        self.done_btn = Button(bottom_frame, text="Save Mapping", command=self.save_mapping, bootstyle="primary")
        self.done_btn.pack(side="left", padx=16)
        self.cancel_btn = Button(bottom_frame, text="Cancel", command=self.cancel, bootstyle="secondary")
        self.cancel_btn.pack(side="left", padx=16)

        self.update_mapping_view()
        self.bind("<Return>", lambda event: self.save_mapping())
        self.bind("<Escape>", lambda event: self.cancel())
        self.update_idletasks()

    def map_selected(self):
        sel1 = self.l1.curselection()
        sel2 = self.l2.curselection()
        if sel1 and sel2 and self.chk_vars1[sel1[0]].get() and self.chk_vars2[sel2[0]].get():
            self.mapdict[int(sel1[0])] = int(sel2[0])
            self.update_mapping_view()

    def unmap_selected(self):
        sel = self.mapping_view.curselection()
        if sel:
            item_idx = sel[0]
            sorted_items = sorted(self.mapdict.items())
            if 0 <= item_idx < len(sorted_items):
                idx1, idx2 = sorted_items[item_idx]
                del self.mapdict[idx1]
                self.update_mapping_view()

    def suggest(self):
        h1 = [h for i, h in enumerate(self.headers1) if self.chk_vars1[i].get()]
        h2 = [h for i, h in enumerate(self.headers2) if self.chk_vars2[i].get()]
        offset1 = [i for i, v in enumerate(self.chk_vars1) if v.get()]
        offset2 = [i for i, v in enumerate(self.chk_vars2) if v.get()]
        smap = suggest_mappings(h1, h2)
        new_mapdict = {}
        for i1, i2 in smap.items():
            new_mapdict[offset1[i1]] = offset2[i2]
        self.mapdict = new_mapdict
        self.update_mapping_view()

    def update_mapping_view(self):
        self.mapping_view.delete(0, tk.END)
        to_remove = [k for k, v in self.mapdict.items()
                     if k >= len(self.headers1) or v >= len(self.headers2) or not self.chk_vars1[k].get() or not self.chk_vars2[v].get()]
        for k in to_remove:
            if k in self.mapdict:
                del self.mapdict[k]
        for idx1, idx2 in sorted(self.mapdict.items()):
            self.mapping_view.insert(tk.END, f'{self.headers1[idx1]}  →  {self.headers2[idx2]}')

    def save_mapping(self):
        self.result = dict(self.mapdict)
        self.result_include1 = [v.get() for v in self.chk_vars1]
        self.result_include2 = [v.get() for v in self.chk_vars2]
        self.destroy()

    def cancel(self):
        self.result = None
        self.result_include1 = None
        self.result_include2 = None
        self.destroy()

    def save_mapping_profile(self):
        fname = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")], title="Save Mapping Profile")
        if fname:
            try:
                data = {
                    "headers1": self.headers1,
                    "headers2": self.headers2,
                    "mapping": self.mapdict,
                    "include1": [v.get() for v in self.chk_vars1],
                    "include2": [v.get() for v in self.chk_vars2],
                }
                with open(fname, "w", encoding="utf-8") as f:
                    json.dump(data, f, indent=2)
                messagebox.showinfo("Success", "Mapping profile saved.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save mapping profile:\n{e}")

    def load_mapping_profile(self):
        fname = filedialog.askopenfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")], title="Load Mapping Profile")
        if fname:
            try:
                with open(fname, "r", encoding="utf-8") as f:
                    data = json.load(f)
                loaded_map = data.get("mapping", {})
                loaded_map = {int(k): int(v) for k, v in loaded_map.items()}
                self.mapdict = loaded_map
                for i, v in enumerate(data.get("include1", [])):
                    if i < len(self.chk_vars1): self.chk_vars1[i].set(v)
                for i, v in enumerate(data.get("include2", [])):
                    if i < len(self.chk_vars2): self.chk_vars2[i].set(v)
                self.update_mapping_view()
                messagebox.showinfo("Loaded", "Mapping profile loaded.\n(check if headers match!)")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load mapping profile:\n{e}")

class ExcelComparatorApp:
    def __init__(self, root):
        self.root = root
        self.settings = {}
        self.load_settings()
        self.style = Style(self.settings.get('theme', 'flatly'))
        self.theme_names = self.style.theme_names()
        self.file1 = ""
        self.file2 = ""
        self.data1 = []
        self.data2 = []
        self.headers1 = []
        self.headers2 = []
        self.sheetnames1 = []
        self.sheetnames2 = []
        self.selected_sheet1 = tk.StringVar()
        self.selected_sheet2 = tk.StringVar()
        self.include1 = []
        self.include2 = []
        self.mapping = mapping_str_to_int(self.settings.get("mapping", {}))
        self.recent_files = self.settings.get("recent_files", [])
        self.recent_outputs = self.settings.get("recent_outputs", [])
        self.recent_filtered_outputs = self.settings.get("recent_filtered_outputs", [])
        self.make_gui()
        self.apply_settings()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def open_mapping(self):
        self.reload_data1()
        self.reload_data2()
        if not self.headers1 or not self.headers2:
            messagebox.showwarning("Mapping", "Load both files and sheets first.")
            return
        dialog = MappingDialog(self.root, self.headers1, self.headers2, self.mapping, self.include1, self.include2)
        self.root.wait_window(dialog)
        if dialog.result is not None:
            self.mapping = mapping_str_to_int(dialog.result)
            self.include1 = dialog.result_include1
            self.include2 = dialog.result_include2
            self.save_settings()
            self.status_var.set("Mapping saved.")

    def load_settings(self):
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, "r") as f:
                self.settings = json.load(f)
        else:
            self.settings = {}
        if "window_geometry" in self.settings:
            try:
                self.root.geometry(self.settings["window_geometry"])
            except Exception:
                pass

    def save_settings(self):
        self.settings["theme"] = self.theme_var.get()
        self.settings["mapping"] = self.mapping
        self.settings["header_font"] = self.header_font.get()
        self.settings["header_size"] = self.header_size.get()
        self.settings["body_font"] = self.body_font.get()
        self.settings["body_size"] = self.body_size.get()
        self.settings["header_fill"] = self.header_fill.get()
        self.settings["header_fontcolor"] = self.header_fontcolor.get()
        self.settings["body_fill"] = self.body_fill.get()
        self.settings["body_fontcolor"] = self.body_fontcolor.get()
        self.settings["match_highlight"] = self.match_highlight.get()
        self.settings["partial_highlight"] = self.partial_highlight.get()
        self.settings["nomatch_highlight"] = self.nomatch_highlight.get()
        self.settings["header_border_thick"] = self.header_border_thick.get()
        self.settings["header_border_color"] = self.header_border_color.get()
        self.settings["body_border_thick"] = self.body_border_thick.get()
        self.settings["body_border_color"] = self.body_border_color.get()
        self.settings["header_height"] = self.header_height.get()
        self.settings["body_height"] = self.body_height.get()
        self.settings["padding"] = self.padding.get()
        self.settings["sort_by_match"] = bool(self.sort_by_match.get())
        self.settings["filtered_output_enabled"] = bool(self.filtered_output_enabled.get())
        self.settings["filtered_output_type"] = self.filtered_output_type.get()
        self.settings["recent_files"] = self.recent_files[-RECENT_LIMIT:]
        self.settings["recent_outputs"] = self.recent_outputs[-RECENT_LIMIT:]
        self.settings["recent_filtered_outputs"] = self.recent_filtered_outputs[-RECENT_LIMIT:]
        self.settings["filtered_output_file"] = self.filtered_output_file_var.get()
        self.settings["export_mapped_only"] = bool(getattr(self, "export_mapped_only", tk.BooleanVar(value=False)).get())
        self.settings["export_match_types_separately"] = bool(self.export_match_types_separately.get())
        self.settings["window_geometry"] = self.root.geometry()
        self.settings["include1"] = self.include1
        self.settings["include2"] = self.include2
        self.settings["selected_sheet1"] = self.selected_sheet1.get()
        self.settings["selected_sheet2"] = self.selected_sheet2.get()
        with open(SETTINGS_FILE, "w") as f:
            json.dump(self.settings, f, indent=2)

    def on_close(self):
        self.save_settings()
        self.root.destroy()

    def change_theme(self, event=None):
        new_theme = self.theme_var.get()
        self.style.theme_use(new_theme)
        self.settings['theme'] = new_theme
        self.save_settings()

    def update_recent_files(self, path):
        if path and path not in self.recent_files:
            self.recent_files.append(path)
            self.recent_files = self.recent_files[-RECENT_LIMIT:]

    def update_recent_outputs(self, path):
        if path and path not in self.recent_outputs:
            self.recent_outputs.append(path)
            self.recent_outputs = self.recent_outputs[-RECENT_LIMIT:]

    def update_recent_filtered_outputs(self, path):
        if path and path not in self.recent_filtered_outputs:
            self.recent_filtered_outputs.append(path)
            self.recent_filtered_outputs = self.recent_filtered_outputs[-RECENT_LIMIT:]

    def make_gui(self):
        self.root.title("Excel Comparator Pro (ttkbootstrap) - Advanced")
        self.root.geometry("1050x820")
        main = Frame(self.root, padding=12)
        main.pack(fill="both", expand=True)

        theme_frame = Frame(main)
        theme_frame.pack(fill="x", pady=(2,0), anchor="ne")
        Label(theme_frame, text="Theme:", font=("Segoe UI", 11, "bold")).pack(side="right", padx=4)
        self.theme_var = tk.StringVar()
        self.theme_combo = Combobox(
            theme_frame,
            textvariable=self.theme_var,
            values=self.theme_names,
            width=16,
            state="readonly"
        )
        self.theme_var.set(self.style.theme.name)
        self.theme_combo.pack(side="right", padx=(4,16))
        self.theme_combo.bind("<<ComboboxSelected>>", self.change_theme)

        banner = Frame(main, bootstyle="primary")
        banner.pack(fill="x", pady=(0,8))
        Label(
            banner, text="Excel Comparator Pro", font=("Segoe UI", 22, "bold"), bootstyle="inverse-primary"
        ).pack(side="left", padx=18, pady=12)
        Label(
            banner, text="with ttkbootstrap", font=("Segoe UI", 10), bootstyle="inverse-primary"
        ).pack(side="right", padx=14, anchor="e")

        files_group = LabelFrame(main, text="1️⃣ Input and Output Files", bootstyle="info", padding=(14, 12))
        files_group.pack(fill="x", padx=2, pady=4)

        Label(files_group, text="File 1:").grid(row=0, column=0, sticky="e", padx=2, pady=4)
        self.f1_var = tk.StringVar()
        self.f1_combo = Combobox(files_group, textvariable=self.f1_var, width=56, values=self.recent_files)
        self.f1_combo.grid(row=0, column=1, padx=2, pady=4)
        Button(files_group, text="Browse", command=lambda: self.pick_file(1), bootstyle="primary-outline").grid(row=0, column=2, padx=2, pady=4)
        Label(files_group, text="Sheet:").grid(row=0, column=3, sticky="e", padx=2)
        self.sheet1_combo = Combobox(files_group, textvariable=self.selected_sheet1, width=18, state="readonly")
        self.sheet1_combo.grid(row=0, column=4, padx=2, pady=4)

        Label(files_group, text="File 2:").grid(row=1, column=0, sticky="e", padx=2, pady=4)
        self.f2_var = tk.StringVar()
        self.f2_combo = Combobox(files_group, textvariable=self.f2_var, width=56, values=self.recent_files)
        self.f2_combo.grid(row=1, column=1, padx=2, pady=4)
        Button(files_group, text="Browse", command=lambda: self.pick_file(2), bootstyle="primary-outline").grid(row=1, column=2, padx=2, pady=4)
        Label(files_group, text="Sheet:").grid(row=1, column=3, sticky="e", padx=2)
        self.sheet2_combo = Combobox(files_group, textvariable=self.selected_sheet2, width=18, state="readonly")
        self.sheet2_combo.grid(row=1, column=4, padx=2, pady=4)

        Label(files_group, text="Output:").grid(row=2, column=0, sticky="e", padx=2, pady=4)
        self.out_var = tk.StringVar()
        self.out_combo = Combobox(files_group, textvariable=self.out_var, width=56, values=self.recent_outputs)
        self.out_combo.grid(row=2, column=1, padx=2, pady=4)
        Button(files_group, text="Browse", command=self.pick_output, bootstyle="success-outline").grid(row=2, column=2, padx=2, pady=4)

        Button(files_group, text="Map Columns", command=self.open_mapping, bootstyle="info-outline").grid(row=0, column=5, rowspan=3, padx=(12, 2), pady=4, sticky="ns")

        self.export_mapped_only = tk.BooleanVar()
        self.export_mapped_only.set(self.settings.get("export_mapped_only", False))
        Checkbutton(files_group,
            text="Export only mapped columns",
            variable=self.export_mapped_only,
            bootstyle="primary-round-toggle"
        ).grid(row=3, column=0, columnspan=6, sticky="w", padx=2, pady=(8,2))

        self.export_match_types_separately = tk.BooleanVar()
        self.export_match_types_separately.set(self.settings.get("export_match_types_separately", False))
        Checkbutton(files_group,
            text="Export Full, Partial, and No Match rows to separate files",
            variable=self.export_match_types_separately,
            bootstyle="info-round-toggle"
        ).grid(row=5, column=0, columnspan=6, sticky="w", padx=2, pady=(4,2))

        partial_frame = Frame(files_group)
        partial_frame.grid(row=6, column=0, columnspan=6, sticky="w", padx=2, pady=(4,2))
        Button(partial_frame, text="Export Partial Match Rows Only", command=self.export_partial_match_rows, bootstyle="warning-outline").pack(side="left", padx=(0,8))
        self.partial_from_var = tk.StringVar()
        self.partial_from_var.set("Both")
        Combobox(partial_frame, textvariable=self.partial_from_var, values=["File1", "File2", "Both"], width=10, state="readonly").pack(side="left", padx=(2,14))

        fmt_group = LabelFrame(main, text="2️⃣ Formatting & Highlighting", bootstyle="warning", padding=(14, 12))
        fmt_group.pack(fill="x", padx=2, pady=8)
        fonts = sorted(font.families())
        Label(fmt_group, text="Header Font:").grid(row=0, column=0, sticky="e", padx=2, pady=(0,2))
        self.header_font = Combobox(fmt_group, values=fonts, width=15)
        self.header_font.grid(row=0, column=1, sticky="w", padx=2, pady=(0,2))
        Label(fmt_group, text="Size:").grid(row=0, column=2, sticky="e")
        self.header_size = Spinbox(fmt_group, from_=8, to=32, width=4)
        self.header_size.grid(row=0, column=3, padx=2)
        Label(fmt_group, text="Fill:").grid(row=0, column=4, sticky="e")
        self.header_fill = Entry(fmt_group, width=10)
        self.header_fill.grid(row=0, column=5, padx=2)
        self.header_fill_swatch = tk.Label(fmt_group, width=2, bg="#f5f1e3", relief="groove")
        self.header_fill_swatch.grid(row=0, column=6, padx=2)
        Button(fmt_group, text="Pick", command=lambda: self.pick_color(self.header_fill, self.header_fill_swatch), bootstyle="secondary").grid(row=0, column=7, padx=2)
        Label(fmt_group, text="Font color:").grid(row=0, column=8, sticky="e")
        self.header_fontcolor = Entry(fmt_group, width=10)
        self.header_fontcolor.grid(row=0, column=9, padx=2)
        self.header_fontcolor_swatch = tk.Label(fmt_group, width=2, bg="#222222", relief="groove")
        self.header_fontcolor_swatch.grid(row=0, column=10, padx=2)
        Button(fmt_group, text="Pick", command=lambda: self.pick_color(self.header_fontcolor, self.header_fontcolor_swatch), bootstyle="secondary").grid(row=0, column=11, padx=2)
        Label(fmt_group, text="Border:").grid(row=0, column=12, sticky="e")
        self.header_border_thick = Spinbox(fmt_group, from_=0, to=3, width=4)
        self.header_border_thick.grid(row=0, column=13, padx=2)
        self.header_border_color = Entry(fmt_group, width=10)
        self.header_border_color.grid(row=0, column=14, padx=2)
        self.header_border_color_swatch = tk.Label(fmt_group, width=2, bg="#333333", relief="groove")
        self.header_border_color_swatch.grid(row=0, column=15, padx=2)
        Button(fmt_group, text="Pick", command=lambda: self.pick_color(self.header_border_color, self.header_border_color_swatch), bootstyle="secondary").grid(row=0, column=16, padx=2)
        Label(fmt_group, text="Body Font:").grid(row=1, column=0, sticky="e", padx=2)
        self.body_font = Combobox(fmt_group, values=fonts, width=15)
        self.body_font.grid(row=1, column=1, sticky="w", padx=2)
        Label(fmt_group, text="Size:").grid(row=1, column=2, sticky="e")
        self.body_size = Spinbox(fmt_group, from_=8, to=32, width=4)
        self.body_size.grid(row=1, column=3, padx=2)
        Label(fmt_group, text="Fill:").grid(row=1, column=4, sticky="e")
        self.body_fill = Entry(fmt_group, width=10)
        self.body_fill.grid(row=1, column=5, padx=2)
        self.body_fill_swatch = tk.Label(fmt_group, width=2, bg="#ffffff", relief="groove")
        self.body_fill_swatch.grid(row=1, column=6, padx=2)
        Button(fmt_group, text="Pick", command=lambda: self.pick_color(self.body_fill, self.body_fill_swatch), bootstyle="secondary").grid(row=1, column=7, padx=2)
        Label(fmt_group, text="Font color:").grid(row=1, column=8, sticky="e")
        self.body_fontcolor = Entry(fmt_group, width=10)
        self.body_fontcolor.grid(row=1, column=9, padx=2)
        self.body_fontcolor_swatch = tk.Label(fmt_group, width=2, bg="#222222", relief="groove")
        self.body_fontcolor_swatch.grid(row=1, column=10, padx=2)
        Button(fmt_group, text="Pick", command=lambda: self.pick_color(self.body_fontcolor, self.body_fontcolor_swatch), bootstyle="secondary").grid(row=1, column=11, padx=2)
        Label(fmt_group, text="Border:").grid(row=1, column=12, sticky="e")
        self.body_border_thick = Spinbox(fmt_group, from_=0, to=3, width=4)
        self.body_border_thick.grid(row=1, column=13, padx=2)
        self.body_border_color = Entry(fmt_group, width=10)
        self.body_border_color.grid(row=1, column=14, padx=2)
        self.body_border_color_swatch = tk.Label(fmt_group, width=2, bg="#aaaaaa", relief="groove")
        self.body_border_color_swatch.grid(row=1, column=15, padx=2)
        Button(fmt_group, text="Pick", command=lambda: self.pick_color(self.body_border_color, self.body_border_color_swatch), bootstyle="secondary").grid(row=1, column=16, padx=2)

        Label(fmt_group, text="Full Match:").grid(row=2, column=0, sticky="e", padx=2, pady=(6,2))
        self.match_highlight = Entry(fmt_group, width=10)
        self.match_highlight.grid(row=2, column=1, padx=2, pady=(6,2))
        self.match_highlight_swatch = tk.Label(fmt_group, width=2, bg="#c6efce", relief="groove")
        self.match_highlight_swatch.grid(row=2, column=2, padx=2, pady=(6,2))
        Button(fmt_group, text="Pick", command=lambda: self.pick_color(self.match_highlight, self.match_highlight_swatch), bootstyle="secondary").grid(row=2, column=3, padx=2, pady=(6,2))
        Label(fmt_group, text="Partial:").grid(row=2, column=4, sticky="e")
        self.partial_highlight = Entry(fmt_group, width=10)
        self.partial_highlight.grid(row=2, column=5, padx=2)
        self.partial_highlight_swatch = tk.Label(fmt_group, width=2, bg="#fff2cc", relief="groove")
        self.partial_highlight_swatch.grid(row=2, column=6, padx=2)
        Button(fmt_group, text="Pick", command=lambda: self.pick_color(self.partial_highlight, self.partial_highlight_swatch), bootstyle="secondary").grid(row=2, column=7, padx=2)
        Label(fmt_group, text="No Match:").grid(row=2, column=8, sticky="e")
        self.nomatch_highlight = Entry(fmt_group, width=10)
        self.nomatch_highlight.grid(row=2, column=9, padx=2)
        self.nomatch_highlight_swatch = tk.Label(fmt_group, width=2, bg="#ffffff", relief="groove")
        self.nomatch_highlight_swatch.grid(row=2, column=10, padx=2)
        Button(fmt_group, text="Pick", command=lambda: self.pick_color(self.nomatch_highlight, self.nomatch_highlight_swatch), bootstyle="secondary").grid(row=2, column=11, padx=2)

        self.sort_by_match = tk.BooleanVar()
        self.sort_by_match.set(self.settings.get("sort_by_match", False))
        self.sort_check = Checkbutton(fmt_group, text="Sort output by match type (Full→Partial→No Match)",
                                      variable=self.sort_by_match, bootstyle="info-round-toggle")
        self.sort_check.grid(row=4, column=0, columnspan=8, pady=(10, 0), sticky="w")

        self.filtered_output_enabled = tk.BooleanVar()
        self.filtered_output_enabled.set(self.settings.get("filtered_output_enabled", False))
        self.filtered_output_type = tk.StringVar()
        self.filtered_output_type.set(self.settings.get("filtered_output_type", "Full Match"))
        self.filtered_output_file_var = tk.StringVar()
        self.filtered_output_file_var.set(self.settings.get("filtered_output_file", ""))

        filter_frame = Frame(fmt_group)
        filter_frame.grid(row=6, column=0, columnspan=17, sticky="w", pady=(12,0))
        self.filter_check = Checkbutton(filter_frame, text="Generate filtered output file", 
                                        variable=self.filtered_output_enabled, bootstyle="success-round-toggle",
                                        command=self.toggle_filtered_output_controls)
        self.filter_check.pack(side="left", padx=(0,10))
        Label(filter_frame, text="Type:").pack(side="left")
        self.filter_type_combo = Combobox(filter_frame, textvariable=self.filtered_output_type, values=["Full Match", "Partial Match", "No Match"], width=15, state="readonly")
        self.filter_type_combo.pack(side="left", padx=(2,10))
        Label(filter_frame, text="File:").pack(side="left")
        self.filter_output_combo = Combobox(filter_frame, textvariable=self.filtered_output_file_var, width=40, values=self.recent_filtered_outputs, state="readonly")
        self.filter_output_combo.pack(side="left", padx=(2,2))
        Button(filter_frame, text="Browse", command=self.pick_filtered_output, bootstyle="secondary").pack(side="left", padx=(2,2))

        row_fmt = Frame(fmt_group)
        row_fmt.grid(row=7, column=0, columnspan=17, pady=(12,4), sticky="w")
        Label(row_fmt, text="Header Row Height:").pack(side="left")
        self.header_height = Spinbox(row_fmt, from_=16, to=80, width=5)
        self.header_height.pack(side="left", padx=(2,10))
        Label(row_fmt, text="Body Row Height:").pack(side="left")
        self.body_height = Spinbox(row_fmt, from_=14, to=80, width=5)
        self.body_height.pack(side="left", padx=(2,10))
        Label(row_fmt, text="Column Padding:").pack(side="left")
        self.padding = Spinbox(row_fmt, from_=0, to=10, width=5)
        self.padding.pack(side="left", padx=2)

        btn_frame = Frame(main)
        btn_frame.pack(fill="x", pady=(18,6))
        Button(btn_frame, text="Compare and Save Output", command=self.compare_and_save, width=30, bootstyle="success").pack(pady=2)

        self.status_var = tk.StringVar(value="Ready.")
        statusbar = Label(main, textvariable=self.status_var, anchor="w", bootstyle="inverse-secondary")
        statusbar.pack(fill="x", side="bottom")

        self.toggle_filtered_output_controls()
        
    def reload_data1(self):
        try:
            wb = openpyxl.load_workbook(self.f1_var.get(), read_only=True)
            ws = wb[self.selected_sheet1.get()]
            self.data1 = [list(row) for row in ws.iter_rows(values_only=True)]
            self.headers1 = list(self.data1[0])
            if not self.include1 or len(self.include1) != len(self.headers1):
                self.include1 = [True]*len(self.headers1)
            wb.close()
        except Exception:
            self.data1 = []
            self.headers1 = []
            self.include1 = []

    def reload_data2(self):
        try:
            wb = openpyxl.load_workbook(self.f2_var.get(), read_only=True)
            ws = wb[self.selected_sheet2.get()]
            self.data2 = [list(row) for row in ws.iter_rows(values_only=True)]
            self.headers2 = list(self.data2[0])
            if not self.include2 or len(self.include2) != len(self.headers2):
                self.include2 = [True]*len(self.headers2)
            wb.close()
        except Exception:
            self.data2 = []
            self.headers2 = []
            self.include2 = []

    def pick_file(self, which):
        fname = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not fname:
            return
        self.update_recent_files(fname)
        wb = openpyxl.load_workbook(fname, read_only=True)
        sheetnames = wb.sheetnames
        wb.close()
        if which == 1:
            self.f1_var.set(fname)
            self.sheetnames1 = sheetnames
            self.sheet1_combo['values'] = self.sheetnames1
            if self.selected_sheet1.get() not in self.sheetnames1:
                self.selected_sheet1.set(self.sheetnames1[0])
            self.reload_data1()
        else:
            self.f2_var.set(fname)
            self.sheetnames2 = sheetnames
            self.sheet2_combo['values'] = self.sheetnames2
            if self.selected_sheet2.get() not in self.sheetnames2:
                self.selected_sheet2.set(self.sheetnames2[0])
            self.reload_data2()
        self.save_settings()
        self.sheet1_combo.bind("<<ComboboxSelected>>", lambda e: self.reload_data1())
        self.sheet2_combo.bind("<<ComboboxSelected>>", lambda e: self.reload_data2())

    def pick_output(self):
        fname = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if fname:
            self.out_var.set(fname)
            self.update_recent_outputs(fname)
            self.out_combo["values"] = self.recent_outputs
            self.save_settings()

    def pick_filtered_output(self):
        fname = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if fname:
            self.filtered_output_file_var.set(fname)
            self.update_recent_filtered_outputs(fname)
            self.filter_output_combo["values"] = self.recent_filtered_outputs
            self.save_settings()

    def pick_color(self, entrybox, swatch):
        col = colorchooser.askcolor(title="Pick Color", color=entrybox.get())[1]
        if col:
            entrybox.delete(0, tk.END)
            entrybox.insert(0, col)
            swatch.config(bg=col)
            self.save_settings()

    def apply_settings(self):
        s = self.settings
        self.recent_files = s.get("recent_files", [])
        self.recent_outputs = s.get("recent_outputs", [])
        self.recent_filtered_outputs = s.get("recent_filtered_outputs", [])
        self.f1_combo["values"] = self.recent_files
        self.f2_combo["values"] = self.recent_files
        self.out_combo["values"] = self.recent_outputs
        self.filter_output_combo["values"] = self.recent_filtered_outputs
        self.header_font.set(s.get("header_font", "Segoe UI"))
        self.header_size.delete(0, tk.END)
        self.header_size.insert(0, s.get("header_size", 13))
        self.header_fill.delete(0, tk.END)
        self.header_fill.insert(0, s.get("header_fill", "#f5f1e3"))
        self.header_fontcolor.delete(0, tk.END)
        self.header_fontcolor.insert(0, s.get("header_fontcolor", "#222222"))
        self.header_fill_swatch.config(bg=s.get("header_fill", "#f5f1e3"))
        self.header_fontcolor_swatch.config(bg=s.get("header_fontcolor", "#222222"))
        self.header_border_thick.delete(0, tk.END)
        self.header_border_thick.insert(0, s.get("header_border_thick", 2))
        self.header_border_color.delete(0, tk.END)
        self.header_border_color.insert(0, s.get("header_border_color", "#333333"))
        self.header_border_color_swatch.config(bg=s.get("header_border_color", "#333333"))
        self.body_font.set(s.get("body_font", "Segoe UI"))
        self.body_size.delete(0, tk.END)
        self.body_size.insert(0, s.get("body_size", 12))
        self.body_fill.delete(0, tk.END)
        self.body_fill.insert(0, s.get("body_fill", "#ffffff"))
        self.body_fontcolor.delete(0, tk.END)
        self.body_fontcolor.insert(0, s.get("body_fontcolor", "#222222"))
        self.body_fill_swatch.config(bg=s.get("body_fill", "#ffffff"))
        self.body_fontcolor_swatch.config(bg=s.get("body_fontcolor", "#222222"))
        self.body_border_thick.delete(0, tk.END)
        self.body_border_thick.insert(0, s.get("body_border_thick", 1))
        self.body_border_color.delete(0, tk.END)
        self.body_border_color.insert(0, s.get("body_border_color", "#aaaaaa"))
        self.body_border_color_swatch.config(bg=s.get("body_border_color", "#aaaaaa"))
        self.match_highlight.delete(0, tk.END)
        self.match_highlight.insert(0, s.get("match_highlight", "#c6efce"))
        self.match_highlight_swatch.config(bg=s.get("match_highlight", "#c6efce"))
        self.partial_highlight.delete(0, tk.END)
        self.partial_highlight.insert(0, s.get("partial_highlight", "#fff2cc"))
        self.partial_highlight_swatch.config(bg=s.get("partial_highlight", "#fff2cc"))
        self.nomatch_highlight.delete(0, tk.END)
        self.nomatch_highlight.insert(0, s.get("nomatch_highlight", "#ffffff"))
        self.nomatch_highlight_swatch.config(bg=s.get("nomatch_highlight", "#ffffff"))
        self.header_height.delete(0, tk.END)
        self.header_height.insert(0, s.get("header_height", 24))
        self.body_height.delete(0, tk.END)
        self.body_height.insert(0, s.get("body_height", 18))
        self.padding.delete(0, tk.END)
        self.padding.insert(0, s.get("padding", 2))
        self.sort_by_match.set(s.get("sort_by_match", False))
        self.filtered_output_enabled.set(s.get("filtered_output_enabled", False))
        self.filtered_output_type.set(s.get("filtered_output_type", "Full Match"))
        self.filtered_output_file_var.set(s.get("filtered_output_file", ""))
        self.export_mapped_only.set(s.get("export_mapped_only", False))
        self.export_match_types_separately.set(s.get("export_match_types_separately", False))
        self.include1 = s.get("include1", [])
        self.include2 = s.get("include2", [])
        self.selected_sheet1.set(s.get("selected_sheet1", ""))
        self.selected_sheet2.set(s.get("selected_sheet2", ""))
        self.theme_var.set(s.get('theme', 'flatly'))
        self.toggle_filtered_output_controls()

    def toggle_filtered_output_controls(self):
        state = "normal" if self.filtered_output_enabled.get() else "disabled"
        self.filter_type_combo.configure(state=state)
        self.filter_output_combo.configure(state=state)

    def show_dashboard(self, counts_A, counts_B, total_A, total_B):
        def pct(val, total):
            return f"{val} ({val/total*100:.1f}%)" if total else "0 (0%)"
        msg = (
            f"Dashboard Summary:\n\n"
            f"File 1 (Rows: {total_A}):\n"
            f"  Full Match: {pct(counts_A['Full Match'], total_A)}\n"
            f"  Partial Match: {pct(counts_A['Partial Match'], total_A)}\n"
            f"  No Match: {pct(counts_A['No Match'], total_A)}\n\n"
            f"File 2 (Rows: {total_B}):\n"
            f"  Full Match: {pct(counts_B['Full Match'], total_B)}\n"
            f"  Partial Match: {pct(counts_B['Partial Match'], total_B)}\n"
            f"  No Match: {pct(counts_B['No Match'], total_B)}"
        )
        messagebox.showinfo("Comparison Summary", msg)

    def get_annotated_rows(self, data_main, data_other, mapping):
        row_info = []
        mapped_indices = list(mapping.items())
        for row_main in data_main[1:]:
            status = "No Match"
            for row_other in data_other[1:]:
                try:
                    matches = [
                        str(row_main[i1]) == str(row_other[i2])
                        for i1, i2 in mapped_indices
                        if i1 < len(row_main) and i2 < len(row_other)
                    ]
                    if matches and all(matches):
                        status = "Full Match"
                        break
                    elif matches and any(matches):
                        status = "Partial Match"
                except Exception:
                    continue
            row_info.append((row_main, status))
        return row_info

    def write_output_sheet(self, ws, row_info, headers, opts, mapping, is_file1, export_mapped_only):
        from openpyxl.styles import Alignment
        hfont = get_font(opts["header_font"], opts["header_size"], True, opts["header_fontcolor"])
        bfont = get_font(opts["body_font"], opts["body_size"], False, opts["body_fontcolor"])
        hfill = get_fill(opts["header_fill"])
        bfill = get_fill(opts["body_fill"])
        hborder = get_border(opts["header_border_thick"], opts["header_border_color"])
        bborder = get_border(opts["body_border_thick"], opts["body_border_color"])
        match_fill = get_fill(opts["match_highlight"])
        partial_fill = get_fill(opts["partial_highlight"])
        nomatch_fill = get_fill(opts["nomatch_highlight"])
        align_center = Alignment(horizontal="center", vertical="center")
        pad = opts["padding"]

        mapped_cols = []
        if export_mapped_only:
            if mapping:
                if is_file1:
                    mapped_cols = [k for k, _ in sorted(mapping.items()) if k < len(headers)]
                else:
                    mapped_cols = [v for _, v in sorted(mapping.items()) if v < len(headers)]
            if not mapped_cols:
                mapped_cols = list(range(len(headers)))
        else:
            mapped_cols = list(range(len(headers)))

        for col_idx, header_index in enumerate(mapped_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=headers[header_index])
            cell.font = hfont
            cell.fill = hfill
            cell.border = hborder
            cell.alignment = align_center
        ws.cell(row=1, column=len(mapped_cols)+1, value="MatchType").font = hfont
        for col_idx in range(1, len(mapped_cols)+2):
            ws.cell(row=1, column=col_idx).fill = hfill
            ws.cell(row=1, column=col_idx).border = hborder
            ws.cell(row=1, column=col_idx).alignment = align_center
        ws.row_dimensions[1].height = opts["header_height"]

        ws.freeze_panes = ws['A2']
        max_row = max(2, len(row_info) + 1)
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(mapped_cols)+1)}{max_row}"

        for idx, (row_main, status) in enumerate(row_info, start=2):
            for col_idx, header_index in enumerate(mapped_cols, 1):
                value = row_main[header_index] if header_index < len(row_main) else ""
                cell = ws.cell(row=idx, column=col_idx, value=value)
                cell.font = bfont
                cell.fill = bfill
                cell.border = bborder
                cell.alignment = align_center

            matchtype_cell = ws.cell(row=idx, column=len(mapped_cols)+1, value=status)
            matchtype_cell.font = bfont
            for col_idx in range(1, len(mapped_cols)+2):
                ws.cell(row=idx, column=col_idx).alignment = align_center
            ws.row_dimensions[idx].height = opts["body_height"]

            if status == "Full Match":
                row_fill = match_fill
                font_bold = False
            elif status == "Partial Match":
                row_fill = partial_fill
                font_bold = True
            else:
                row_fill = nomatch_fill
                font_bold = True
            for col_idx in range(1, len(mapped_cols)+2):
                ws.cell(row=idx, column=col_idx).fill = row_fill
                if font_bold:
                    ws.cell(row=idx, column=col_idx).font = get_font(
                        opts["body_font"],
                        opts["body_size"],
                        True,
                        opts["body_fontcolor"]
                    )
        autofit_columns(ws, extra_padding=pad)

    def compare_and_save(self):
        self.save_settings()
        self.reload_data1()
        self.reload_data2()
        opts = {
            "header_font": self.header_font.get(),
            "header_size": int(self.header_size.get()),
            "header_fill": self.header_fill.get(),
            "header_fontcolor": self.header_fontcolor.get(),
            "header_border_thick": int(self.header_border_thick.get()),
            "header_border_color": self.header_border_color.get(),
            "body_font": self.body_font.get(),
            "body_size": int(self.body_size.get()),
            "body_fill": self.body_fill.get(),
            "body_fontcolor": self.body_fontcolor.get(),
            "body_border_thick": int(self.body_border_thick.get()),
            "body_border_color": self.body_border_color.get(),
            "match_highlight": self.match_highlight.get(),
            "partial_highlight": self.partial_highlight.get(),
            "nomatch_highlight": self.nomatch_highlight.get(),
            "header_height": int(self.header_height.get()),
            "body_height": int(self.body_height.get()),
            "padding": int(self.padding.get())
        }
        try:
            headers1, headers2 = self.headers1, self.headers2
            data1, data2 = self.data1, self.data2
            include1 = [i for i, v in enumerate(self.include1) if v]
            include2 = [i for i, v in enumerate(self.include2) if v]
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open files:\n{e}")
            return

        outname = self.out_var.get()
        if not outname:
            outname = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if not outname:
                return
            self.out_var.set(outname)
            self.update_recent_outputs(outname)
            self.out_combo["values"] = self.recent_outputs

        mapping = {k: v for k, v in self.mapping.items() if k in include1 and v in include2}
        used_headers1 = [headers1[i] for i in include1]
        used_headers2 = [headers2[i] for i in include2]
        used_data1 = [[row[i] for i in include1] for row in data1]
        used_data2 = [[row[i] for i in include2] for row in data2]

        row_info_A = self.get_annotated_rows(used_data1, used_data2, mapping)
        reverse_mapping = {v: k for k, v in mapping.items()}
        row_info_B = self.get_annotated_rows(used_data2, used_data1, reverse_mapping)

        if self.sort_by_match.get():
            sort_order = {"Full Match": 0, "Partial Match": 1, "No Match": 2}
            row_info_A.sort(key=lambda x: sort_order.get(x[1], 99))
            row_info_B.sort(key=lambda x: sort_order.get(x[1], 99))

        def count_types(row_info):
            d = {"Full Match": 0, "Partial Match": 0, "No Match": 0}
            for _, status in row_info:
                if status in d: d[status] += 1
            return d
        counts_A = count_types(row_info_A)
        counts_B = count_types(row_info_B)
        total_A = len(row_info_A)
        total_B = len(row_info_B)
        self.show_dashboard(counts_A, counts_B, total_A, total_B)

        outwb = openpyxl.Workbook()
        wsA = outwb.active
        wsA.title = "File1"
        wsB = outwb.create_sheet("File2")
        self.write_output_sheet(
            wsA, row_info_A, used_headers1, opts, mapping, is_file1=True, 
            export_mapped_only=self.export_mapped_only.get()
        )
        self.write_output_sheet(
            wsB, row_info_B, used_headers2, opts, reverse_mapping, is_file1=False, 
            export_mapped_only=self.export_mapped_only.get()
        )
        outwb.save(outname)
        self.status_var.set(f"Output saved: {outname}")
        self.update_recent_outputs(outname)
        self.out_combo["values"] = self.recent_outputs

        if self.export_match_types_separately.get():
            base, ext = os.path.splitext(outname)
            match_types = ["Full Match", "Partial Match", "No Match"]
            for mt in match_types:
                for which, row_info, headers, mapping, is_file1 in [
                    ("file1", row_info_A, used_headers1, mapping, True),
                    ("file2", row_info_B, used_headers2, reverse_mapping, False)
                ]:
                    rows = [row for row in row_info if row[1] == mt]
                    if not rows: continue
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = which.capitalize()
                    self.write_output_sheet(
                        ws, rows, headers, opts, mapping, is_file1=is_file1,
                        export_mapped_only=self.export_mapped_only.get()
                    )
                    fname = f"{base}_{which}_{mt.replace(' ', '').lower()}{ext}"
                    wb.save(fname)
            messagebox.showinfo("Exported", "Separate files for each match type have been saved in the output directory.")

        if self.filtered_output_enabled.get():
            filter_type = self.filtered_output_type.get()
            filtered_outname = self.filtered_output_file_var.get()
            if not filtered_outname:
                filtered_outname = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
                if not filtered_outname:
                    self.status_var.set("Filtered output not saved (file not selected).")
                    return
                self.filtered_output_file_var.set(filtered_outname)
                self.update_recent_filtered_outputs(filtered_outname)
                self.filter_output_combo["values"] = self.recent_filtered_outputs
                self.save_settings()
            filtered_rows_A = [row for row in row_info_A if row[1] == filter_type]
            filtered_rows_B = [row for row in row_info_B if row[1] == filter_type]
            filtered_wb = openpyxl.Workbook()
            fwsA = filtered_wb.active
            fwsA.title = "File1"
            fwsB = filtered_wb.create_sheet("File2")
            self.write_output_sheet(
                fwsA, filtered_rows_A, used_headers1, opts, mapping, is_file1=True, 
                export_mapped_only=self.export_mapped_only.get()
            )
            self.write_output_sheet(
                fwsB, filtered_rows_B, used_headers2, opts, reverse_mapping, is_file1=False, 
                export_mapped_only=self.export_mapped_only.get()
            )
            filtered_wb.save(filtered_outname)
            self.status_var.set(f"Filtered output saved: {filtered_outname}")
            self.update_recent_filtered_outputs(filtered_outname)
            self.filter_output_combo["values"] = self.recent_filtered_outputs

        self.save_settings()
        messagebox.showinfo("Saved", "Output files saved successfully.")

    def export_partial_match_rows(self):
        self.reload_data1()
        self.reload_data2()
        try:
            headers1, headers2 = self.headers1, self.headers2
            data1, data2 = self.data1, self.data2
            include1 = [i for i, v in enumerate(self.include1) if v]
            include2 = [i for i, v in enumerate(self.include2) if v]
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open files:\n{e}")
            return

        mapping = {k: v for k, v in self.mapping.items() if k in include1 and v in include2}
        reverse_mapping = {v: k for k, v in mapping.items()}
        used_headers1 = [headers1[i] for i in include1]
        used_headers2 = [headers2[i] for i in include2]
        used_data1 = [[row[i] for i in include1] for row in data1]
        used_data2 = [[row[i] for i in include2] for row in data2]
        row_info_A = self.get_annotated_rows(used_data1, used_data2, mapping)
        row_info_B = self.get_annotated_rows(used_data2, used_data1, reverse_mapping)

        from_opt = self.partial_from_var.get()
        outname = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Export Partial Match Rows")
        if not outname:
            return

        opts = {
            "header_font": self.header_font.get(),
            "header_size": int(self.header_size.get()),
            "header_fill": self.header_fill.get(),
            "header_fontcolor": self.header_fontcolor.get(),
            "header_border_thick": int(self.header_border_thick.get()),
            "header_border_color": self.header_border_color.get(),
            "body_font": self.body_font.get(),
            "body_size": int(self.body_size.get()),
            "body_fill": self.body_fill.get(),
            "body_fontcolor": self.body_fontcolor.get(),
            "body_border_thick": int(self.body_border_thick.get()),
            "body_border_color": self.body_border_color.get(),
            "match_highlight": self.match_highlight.get(),
            "partial_highlight": self.partial_highlight.get(),
            "nomatch_highlight": self.nomatch_highlight.get(),
            "header_height": int(self.header_height.get()),
            "body_height": int(self.body_height.get()),
            "padding": int(self.padding.get())
        }
        export_mapped_only = self.export_mapped_only.get()
        outwb = openpyxl.Workbook()

        if from_opt in ("File1", "Both"):
            wsA = outwb.active
            wsA.title = "File1"
            partial_A = [row for row in row_info_A if row[1] == "Partial Match"]
            self.write_output_sheet(wsA, partial_A, used_headers1, opts, mapping, is_file1=True, export_mapped_only=export_mapped_only)
        if from_opt in ("File2", "Both"):
            if from_opt == "Both":
                wsB = outwb.create_sheet("File2")
            else:
                wsB = outwb.active
                wsB.title = "File2"
            partial_B = [row for row in row_info_B if row[1] == "Partial Match"]
            self.write_output_sheet(wsB, partial_B, used_headers2, opts, reverse_mapping, is_file1=False, export_mapped_only=export_mapped_only)

        outwb.save(outname)
        self.update_recent_outputs(outname)
        messagebox.showinfo("Exported", f"Partial match rows exported to:\n{outname}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelComparatorApp(root)
    root.mainloop()
